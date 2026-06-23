# =========================
# app.py
# =========================

from flask import Flask, request, redirect, url_for, render_template, flash, jsonify,send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import (db,Usuario, Role, Permission,Module,Proyecto,Vehiculo,Alerta,Operador,Kardex,Rendimiento,Tanque)
from openpyxl.styles import PatternFill
from config import Config
from functools import wraps
from flask import jsonify
import os
from werkzeug.utils import secure_filename
from flask import get_flashed_messages
import re
import json
import uuid
from sqlalchemy import cast, String,or_,func
from datetime import datetime,timedelta
from sqlalchemy import text
from openpyxl import Workbook
import io
import pytz

PERU_TZ = pytz.timezone("America/Lima")

def now_lima():
    return datetime.now(PERU_TZ).replace(tzinfo=None)


# ------------------------
# APP
# ------------------------
app = Flask(__name__, static_folder='static')
app.config.from_object(Config)


# ------------------------
# INIT
# ------------------------
db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = None


@login_manager.unauthorized_handler
def unauthorized():
    return redirect(url_for('login'))


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Usuario, int(user_id))


# ------------------------
# CONFIG PERMISOS
# ------------------------
MODULOS = [
    "usuarios",
    "roles",
    "proyectos",
    "vehiculos",
    "operadores",
    "tanques",
    "rendimientos",
    "kardex"
]

ACCIONES = ["ver", "crear", "editar", "eliminar"]

ACCIONES_KARDEX = ["ver","crear", "editar","eliminar"]


def now_utc():
    return datetime.utcnow()
# ------------------------
# SEED (MANUAL)
# ------------------------
def seed_data():
    try:
        role_admin = Role.query.filter_by(name="admin").first()
        if not role_admin:
            role_admin = Role(name="admin")
            db.session.add(role_admin)
            db.session.commit()

        admin = Usuario.query.filter_by(username="admin").first()
        if not admin:
            admin = Usuario(
                username="admin",
                email="admin@test.com",
                full_name="Administrador"
            )
            admin.set_password("43737510")
            db.session.add(admin)
            db.session.commit()

        if role_admin not in admin.roles:
            admin.roles.append(role_admin)
            db.session.commit()

        for nombre_modulo in MODULOS:
            mod = Module.query.filter_by(name=nombre_modulo).first()

            if not mod:
                mod = Module(name=nombre_modulo)
                db.session.add(mod)
                db.session.commit()

            acciones_modulo = ACCIONES_KARDEX if nombre_modulo == "kardex" else ACCIONES

            for acc in acciones_modulo:
                perm = Permission.query.filter_by(
                    module_id=mod.id,
                    action=acc
                ).first()

                if not perm:
                    db.session.add(Permission(module_id=mod.id, action=acc))

        db.session.commit()

        for perm in Permission.query.all():
            if perm not in role_admin.permissions:
                role_admin.permissions.append(perm)

        db.session.commit()

        print("✅ Datos iniciales OK")

    except Exception as e:
        print("⚠️ Error en seed:", e)


# ------------------------
# SOLO TEST DE ARRANQUE (LIVIANO)
# ------------------------
#  https://tu-app.onrender.com/init para ejecutar manualmente el init (seed_data())
@app.route("/init")
def init():
    try:
        seed_data()
        return {"msg": "seed ejecutado"}, 200
    except Exception as e:
        return {"error": str(e)}, 500

# ------------------------
# CONTEXTO INICIALIZACION (SEGURO)
# ------------------------
with app.app_context():
    try:
        print("🚀 Iniciando aplicación...")

        # SOLO crear tablas
        db.create_all()

        print("✅ Tablas verificadas")

    except Exception as e:
        print("❌ Error al iniciar BD:", e)

#-- Funciones globales--
def validar_texto(valor, campo, min_len=3, max_len=150):
    if not valor:
        return f"{campo} es obligatorio"
    if len(valor) < min_len:
        return f"{campo} debe tener al menos {min_len} caracteres"
    if len(valor) > max_len:
        return f"{campo} no debe exceder {max_len} caracteres"
    return None


def validar_numero(valor, campo, tipo=float, minimo=0):
    try:
        num = tipo(valor)
        if num < minimo:
            return f"{campo} no puede ser negativo"
        return None
    except:
        return f"{campo} inválido"


def validar_email(email):
    if not email:
        return None
    regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    if not re.match(regex, email):
        return "Email inválido"
    return None

# =========================
# PERMISOS
# =========================
def permission_required(modulo, accion):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("login"))

            permiso = f"{modulo}.{accion}"

            if not current_user.has_permission(permiso):
                flash("No tienes permisos", "danger")
                return redirect(url_for("index"))

            return f(*args, **kwargs)
        return wrapper
    return decorator


# =========================
# RUTAS BASE
# =========================

@app.route('/')
@login_required
def index():
    return redirect(url_for("dashboard_gerencial"))
    #return redirect(url_for("rendimiento_list"))
    
    

@app.route('/login', methods=['GET','POST'])
def login():

    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == 'POST':
        usuario = Usuario.query.filter_by(username=request.form['username']).first()

        if usuario and usuario.check_password(request.form['password']):
            login_user(usuario)
            return redirect(url_for('index'))

        flash('Credenciales inválidas','danger')

    return render_template("login.html")

@app.route('/logout')
@login_required
def logout():
    from flask import session
    logout_user()
    session.clear()  # 🔥 borra toda la sesión
    return redirect(url_for('login'))


# =========================
# CRUD USUARIOS
# =========================
@app.route("/usuarios")
@login_required
@permission_required("usuarios","ver")
def usuarios_list():
    #print([ (p.module.name, p.action) for r in current_user.roles for p in r.permissions ])
    #return "SI FUNCIONA USUARIOS"
    return render_template(
        "usuarios.html",
        lista=Usuario.query.all(),
        roles=Role.query.all()
    )
@app.route("/usuarios/nuevo", methods=["POST"])
@login_required
@permission_required("usuarios","crear")
def usuarios_nuevo():
    try:
        # -------------------------
        # OBTENER DATOS
        # -------------------------
        username = request.form.get("username")
        password = request.form.get("password")
        email = request.form.get("email")
        full_name = request.form.get("full_name")
        roles_ids = request.form.getlist("roles")  # 👈 importante
        proyecto_ids=request.form.getlist("proyectos")

        # -------------------------
        # VALIDACIONES
        # -------------------------
        if not username or not password:
            flash("Usuario y contraseña son obligatorios", "danger")
            return redirect(url_for("usuarios_list"))

        # Usuario único
        if Usuario.query.filter_by(username=username).first():
            flash("El username ya existe", "warning")
            return redirect(url_for("usuarios_list"))

        # Email único (opcional pero recomendado)
        if email and Usuario.query.filter_by(email=email).first():
            flash("El email ya está registrado", "warning")
            return redirect(url_for("usuarios_list"))

        # -------------------------
        # CREAR USUARIO
        # -------------------------
        nuevo_usuario = Usuario(
            username=username,
            email=email,
            full_name=full_name
        )

        # 🔐 usar tu método del modelo
        nuevo_usuario.set_password(password)

        # -------------------------
        # ASIGNAR ROLES
        # -------------------------
        if roles_ids:
            roles = Role.query.filter(Role.id.in_(roles_ids)).all()
            nuevo_usuario.roles = roles  # 👈 relación many-to-many
        # -------------------------
        # ASIGNAR Proyectos
        # -------------------------
        if proyecto_ids:
            proyectos = Proyecto.query.filter(Proyecto.id.in_(proyecto_ids)).all()
            nuevo_usuario.proyectos = proyectos 
        # -------------------------
        # GUARDAR
        # -------------------------
        db.session.add(nuevo_usuario)
        db.session.commit()

        flash("Usuario creado correctamente", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error al crear usuario: {str(e)}", "danger")

    return redirect(url_for("usuarios_list"))

@app.route("/usuarios/editar/<int:id>", methods=["POST"])
@login_required
@permission_required("usuarios","editar")
def usuarios_editar(id):
    usuario = Usuario.query.get_or_404(id)

    try:
        username = request.form.get("username")
        email = request.form.get("email")
        full_name = request.form.get("full_name")
        roles_ids = request.form.getlist("roles")
        proyectos_ids=request.form.getlist("proyectos")

        # -------------------------
        # VALIDACIONES
        # -------------------------
        if not username:
            flash("El username es obligatorio", "danger")
            return redirect(url_for("usuarios_list"))

        # Validar username único (excepto el mismo usuario)
        existe = Usuario.query.filter(Usuario.username == username, Usuario.id != id).first()
        if existe:
            flash("El username ya está en uso", "warning")
            return redirect(url_for("usuarios_list"))

        # Validar email único
        if email:
            existe_email = Usuario.query.filter(Usuario.email == email, Usuario.id != id).first()
            if existe_email:
                flash("El email ya está en uso", "warning")
                return redirect(url_for("usuarios_list"))

        # -------------------------
        # ACTUALIZAR DATOS
        # -------------------------
        usuario.username = username
        usuario.email = email
        usuario.full_name = full_name

        # -------------------------
        # ACTUALIZAR ROLES
        # -------------------------
        usuario.roles = []  # limpiar roles actuales

        if roles_ids:
            roles = Role.query.filter(Role.id.in_(roles_ids)).all()
            usuario.roles = roles
        # -------------------------
        # ACTUALIZAR proyectos
        # -------------------------
        if proyectos_ids:
            proyectos = Proyecto.query.filter(Proyecto.id.in_(proyectos_ids)).all()
            usuario.proyectos = proyectos    

        db.session.commit()
        flash("Usuario actualizado correctamente", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error al actualizar: {str(e)}", "danger")

    return redirect(url_for("usuarios_list"))

@app.route("/usuarios/eliminar/<int:id>", methods=["POST"])
@login_required
@permission_required("usuarios","eliminar")
def usuarios_eliminar(id):
    usuario = Usuario.query.get_or_404(id)

    try:
        # ⚠️ evitar eliminarse a sí mismo
        if usuario.id == current_user.id:
            flash("No puedes eliminar tu propio usuario", "warning")
            return redirect(url_for("usuarios_list"))

        db.session.delete(usuario)
        db.session.commit()

        flash("Usuario eliminado correctamente", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar: {str(e)}", "danger")

    return redirect(url_for("usuarios_list"))

# =========================
# CRUD ROLES
# =========================

@app.route("/roles")
@login_required
@permission_required("roles", "ver")
def roles_list():
    
    permisos = Permission.query.all()

    from collections import defaultdict
    permisos_agrupados = defaultdict(list)

    for p in permisos:
        modulo = p.module.name if p.module else "Sin módulo"
        permisos_agrupados[modulo].append(p)

    return render_template(
        "roles.html",
        lista=Role.query.all(),
        permisos=permisos,  # reutilizas
        permisos_agrupados=permisos_agrupados,
        modulos=Module.query.all()
    )


@app.route("/roles/nuevo", methods=["POST"])
@login_required
@permission_required("roles", "crear")
def roles_nuevo():
    try:
        name = request.form.get("name")
        permisos_ids = request.form.getlist("permisos")

        if not name:
            flash("El nombre del rol es obligatorio", "danger")
            return redirect(url_for("roles_list"))

        if Role.query.filter_by(name=name).first():
            flash("El rol ya existe", "warning")
            return redirect(url_for("roles_list"))

        nuevo = Role(name=name)

        if permisos_ids:
            permisos = Permission.query.filter(Permission.id.in_(permisos_ids)).all()
            nuevo.permissions = permisos

        db.session.add(nuevo)
        db.session.commit()

        flash("Rol creado correctamente", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("roles_list"))


@app.route("/roles/editar/<int:id>", methods=["POST"])
@login_required
@permission_required("roles", "editar")
def roles_editar(id):
    rol = Role.query.get_or_404(id)

    try:
        rol.name = request.form.get("name")
        permisos_ids = request.form.getlist("permisos")

        # actualizar permisos
        rol.permissions = []
        if permisos_ids:
            permisos = Permission.query.filter(Permission.id.in_(permisos_ids)).all()
            rol.permissions = permisos

        db.session.commit()
        flash("Rol actualizado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("roles_list"))


@app.route("/roles/eliminar/<int:id>", methods=["POST"])
@login_required
@permission_required("roles", "eliminar")
def roles_eliminar(id):
    rol = Role.query.get_or_404(id)

    try:
        db.session.delete(rol)
        db.session.commit()
        flash("Rol eliminado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"No se pudo eliminar: {str(e)}", "danger")

    return redirect(url_for("roles_list"))


# =========================
# VEHICULOS
# =========================
@app.route("/vehiculos")
@login_required
@permission_required("vehiculos", "ver")
def vehiculos_list():
    return render_template(
        "vehiculos.html",
        lista=Vehiculo.query.all(),
        proyectos=Proyecto.query.all()
    )


@app.route("/vehiculos/nuevo", methods=["POST"])
@login_required
@permission_required("vehiculos", "crear")
def vehiculos_nuevo():

    try:
        nombre = request.form.get("nombre", "").strip()
        placa = request.form.get("placa", "").strip()
        tipo = request.form.get("tipo")
        proyecto_id = request.form.get("proyecto_id")
        rendimiento = request.form.get("rendimiento_promedio")

        if not nombre:
            flash("Debe ingresar el equipo", "danger")
            return redirect(url_for("vehiculos_list"))
        
        # Validar vehiculo único (excepto el mismo usuario)
        existe = Vehiculo.query.filter(Vehiculo.placa == placa).first()
        if existe:
            flash("La placa del vehículo ya está en uso", "warning")
            return redirect(url_for("vehiculos_list"))
        
        try:
            rendimiento = float(rendimiento) if rendimiento else None
        except:
            flash("Rendimiento inválido", "danger")
            return redirect(url_for("vehiculos_list"))
        

        nuevo = Vehiculo(
            nombre=nombre,
            placa=placa,
            tipo=tipo,
            proyecto_id=int(proyecto_id) if proyecto_id else None,
            rendimiento_promedio=rendimiento
        )

        db.session.add(nuevo)
        db.session.commit()

        flash("✅ Vehículo creado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("vehiculos_list"))


@app.route("/vehiculos/editar/<int:id>", methods=["POST"])
@login_required
@permission_required("vehiculos", "editar")
def vehiculos_editar(id):
    vehiculo = Vehiculo.query.get_or_404(id)

    try:
        nombre = request.form.get("nombre")
        placa = request.form.get("placa")
        tipo = request.form.get("tipo")
        rendimiento = request.form.get("rendimiento_promedio")
        proyecto_id=request.form.get("proyecto_id")

        # -------------------------
        # VALIDACIONES
        # -------------------------
        if not nombre:
            flash("El nombre es obligatorio", "danger")
            return redirect(url_for("vehiculos_list"))

        # -------------------------
        # ACTUALIZAR DATOS
        # -------------------------
        vehiculo.nombre = nombre
        vehiculo.placa = placa
        vehiculo.tipo = tipo
        vehiculo.rendimiento_promedio=float(rendimiento) if rendimiento else None
        vehiculo.proyecto_id=int(proyecto_id) if proyecto_id else None

        db.session.commit()
        flash("Vehiculo actualizado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("vehiculos_list"))


@app.route("/vehiculo/eliminar/<int:id>", methods=["POST"])
@login_required
@permission_required("vehiculos", "eliminar")
def vehiculos_eliminar(id):

    vehiculo = Vehiculo.query.get_or_404(id)

    try:
        db.session.delete(vehiculo)
        db.session.commit()
        flash("Vehiculo eliminado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"No se pudo eliminar: {str(e)}", "danger")
        flash(str(e), "danger")

    return redirect(url_for("vehiculos_list"))

# =========================
# PROYECTOS
# =========================

@app.route("/proyectos")
@login_required
@permission_required("proyectos","ver")
def proyectos_list():

    lista = Proyecto.query.order_by(Proyecto.id.desc()).all()

    return render_template(
        "proyectos.html",
        lista=lista
    )

@app.route("/proyectos/nuevo", methods=["POST"])
@login_required
@permission_required("proyectos","crear")
def proyectos_nuevo():

    nombre = request.form.get("nombre")
    ubicacion = request.form.get("ubicacion")
    nombre_corto = request.form.get("nombre_corto")

    nuevo = Proyecto(
        nombre=nombre,
        ubicacion=ubicacion,
        nombre_corto=nombre_corto,
        activo=True
    )

    db.session.add(nuevo)
    db.session.commit()

    flash("Proyecto registrado correctamente", "success")

    return redirect(url_for("proyectos_list"))


@app.route("/proyectos/editar/<int:id>", methods=["POST"])
@login_required
@permission_required("proyectos","editar")
def proyectos_editar(id):

    proyecto = Proyecto.query.get_or_404(id)

    proyecto.nombre = request.form.get("nombre")
    proyecto.ubicacion = request.form.get("ubicacion")
    proyecto.nombre_corto = request.form.get("nombre_corto")
    proyecto.activo = True if request.form.get("activo") == "1" else False

    db.session.commit()

    flash("Proyecto actualizado correctamente", "success")

    return redirect(url_for("proyectos_list"))


@app.route("/proyectos/eliminar/<int:id>", methods=["POST"])
@login_required
@permission_required("proyectos","eliminar")
def proyectos_eliminar(id):

    proyecto = Proyecto.query.get_or_404(id)

    db.session.delete(proyecto)
    db.session.commit()

    flash("Proyecto eliminado correctamente", "success")

    return redirect(url_for("proyectos_list"))

# =========================
# OPERADORES
# =========================
@app.route("/operadores")
@login_required
@permission_required("operadores", "ver")
def operadores_list():
    return render_template(
        "operadores.html",
        lista=Operador.query.all(),
        proyectos=Proyecto.query.all()
    )

@app.route("/operadores/nuevo", methods=["POST"])
@login_required
@permission_required("operadores", "crear")
def operadores_nuevo():

    try:
        nombre = request.form.get("nombre", "").strip()
        documento = request.form.get("documento", "").strip()
        proyecto_id = request.form.get("proyecto_id")

        if not nombre:
            flash("Debe ingresar nombre", "danger")
            return redirect(url_for("operadores_list"))

        nuevo = Operador(
            nombre=nombre,
            documento=documento,
            proyecto_id=int(proyecto_id) if proyecto_id else None
        )

        db.session.add(nuevo)
        db.session.commit()

        flash("✅ Operador creado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("operadores_list"))


@app.route("/operadores/editar/<int:id>", methods=["POST"])
@login_required
@permission_required("operadores", "editar")
def operadores_editar(id):

    operador = Operador.query.get_or_404(id)

    try:
        nombre = request.form.get("nombre", "").strip()
        documento = request.form.get("documento", "").strip()
        proyecto_id = request.form.get("proyecto_id")

        if not nombre:
            flash("Debe ingresar nombre", "danger")
            return redirect(url_for("operadores_list"))

        operador.nombre = nombre
        operador.documento = documento
        operador.proyecto_id = int(proyecto_id) if proyecto_id else None

        db.session.commit()

        flash("Operador actualizado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("operadores_list"))


@app.route("/operadores/eliminar/<int:id>", methods=["POST"])
@login_required
@permission_required("operadores", "eliminar")
def operadores_eliminar(id):

    operador = Operador.query.get_or_404(id)

    try:
        db.session.delete(operador)
        db.session.commit()

        flash("Operador eliminado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("operadores_list"))
# =========================
# TANQUES
# =========================

@app.route("/tanques")
@login_required
@permission_required("tanques", "ver")
def tanques_list():
    return render_template(
        "tanques.html",
        lista=Tanque.query.all(),
        proyectos=Proyecto.query.all()
    )

@app.route("/tanques/nuevo", methods=["POST"])
@login_required
@permission_required("tanques", "crear")
def tanques_nuevo():

    try:
        nombre = request.form.get("nombre", "").strip()
        capacidad = request.form.get("capacidad")
        stock = request.form.get("stock_actual")
        minimo = request.form.get("stock_minimo")
        proyecto_id = request.form.get("proyecto_id")

        if not nombre:
            flash("Debe ingresar nombre", "danger")
            return redirect(url_for("tanques_list"))

        try:
            capacidad = float(capacidad)
            minimo = float(minimo)
        except:
            flash("Valores numéricos inválidos", "danger")
            return redirect(url_for("tanques_list"))

        nuevo = Tanque(
            nombre=nombre,
            capacidad=capacidad,
            stock_actual=stock,
            stock_minimo=minimo,
            proyecto_id=int(proyecto_id) if proyecto_id else None,
        )

        db.session.add(nuevo)
        db.session.commit()

        flash("✅ Tanque creado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("tanques_list"))

@app.route("/tanques/editar/<int:id>", methods=["POST"])
@login_required
@permission_required("tanques", "editar")
def tanques_editar(id):

    tanque = Tanque.query.get_or_404(id)

    try:

        tanque.nombre = request.form.get("nombre")
        tanque.proyecto_id = int(
            request.form.get("proyecto_id")
        )

        tanque.capacidad = float(
            request.form.get("capacidad") or 0
        )

        tanque.stock_actual = float(
            request.form.get("stock_actual") or 0
        )

        tanque.stock_minimo = float(
            request.form.get("stock_minimo") or 0
        )

        db.session.commit()

        flash(
            "Tanque actualizado",
            "success"
        )

    except Exception as e:

        db.session.rollback()

        flash(
            f"Error: {str(e)}",
            "danger"
        )

    return redirect(
        url_for("tanques_list")
    )

@app.route("/tanques/eliminar/<int:id>", methods=["POST"])
@login_required
@permission_required("tanques", "eliminar")
def tanques_eliminar(id):

    tanque = Tanque.query.get_or_404(id)

    try:
        db.session.delete(tanque)
        db.session.commit()

        flash("Tanque eliminado", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("tanques_list"))
#--------------------------------------------------
# KARDEX
#--------------------------------------------------
@app.route("/kardex")
@login_required
@permission_required("kardex", "ver")
def kardex_list():
    query = Kardex.query

    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    proyecto_id = request.args.get("proyecto_id")
    vehiculo_id = request.args.get("vehiculo_id")
    tipo = request.args.get("tipo")
    buscar = request.args.get("buscar")

    if fecha_inicio:
        query = query.filter(
            Kardex.fecha >= fecha_inicio
        )

    if fecha_fin:
        query = query.filter(
            Kardex.fecha < (
                datetime.strptime(
                    fecha_fin,
                    "%Y-%m-%d"
                ) + timedelta(days=1)
            )
        )

    if proyecto_id:
        query = query.filter(
            Kardex.proyecto_id == proyecto_id
        )

    if vehiculo_id:
        query = query.filter(
            Kardex.vehiculo_id == vehiculo_id
        )

    if tipo:
        query = query.filter(
            Kardex.tipo == tipo
        )

    if buscar:
        query = query.filter(
            db.or_(
                Kardex.parte_diario.ilike(f"%{buscar}%"),
                Kardex.factura.ilike(f"%{buscar}%"),
                Kardex.observacion.ilike(f"%{buscar}%")
            )
        )
    lista = query.order_by(
        Kardex.fecha.desc(),
        Kardex.activo == True
    ).all()

    fecha_actual = now_lima().strftime("%Y-%m-%dT%H:%M")
    return render_template(
        "kardex.html",
        lista=lista,#Kardex.query.filter(Kardex.activo == True).order_by(Kardex.fecha.desc()).all(),
        proyectos=Proyecto.query.filter_by(activo=True).all(),
        vehiculos=Vehiculo.query.all(),
        tanques=Tanque.query.all(),
        operadores=Operador.query.all(),
        fecha_actual=fecha_actual
    )


@app.route("/kardex/nuevo", methods=["POST"])
@login_required
@permission_required("kardex", "crear")
def kardex_nuevo():

    try:
        proyecto_id=request.form.get("proyecto_id")
        tipo = request.form.get("tipo")
        tanque_id = request.form.get("tanque_id")
        vehiculo_id = request.form.get("vehiculo_id")
        operador_id = request.form.get("operador_id")
        tanque_lleno = request.form.get("tanque_lleno") == "on"
        
        precio_unitario = request.form.get("precio_unitario")
        proveedor = request.form.get("proveedor")
        factura = request.form.get("factura")


        fecha_str = request.form.get("fecha")
        cantidad = request.form.get("cantidad")
        h_final = request.form.get("horometro_final")

        hace_5_seg = now_utc() - timedelta(seconds=5)

        existe = Kardex.query.filter(
            Kardex.vehiculo_id == vehiculo_id,
            Kardex.cantidad == cantidad,
            Kardex.fecha >= hace_5_seg,
            Kardex.activo == True
        ).first()

        if existe:
            flash("Movimiento duplicado detectado", "warning")
            return redirect(url_for("kardex_list"))

        # =========================
        # PARSEO SEGURO
        # =========================
        print(precio_unitario)
        print(cantidad)
        try:
            precio_unitario = float(precio_unitario) if precio_unitario else 0.0
        except:
            flash("Precio inválido", "danger")
            return redirect(url_for("kardex_list"))

        try:
            cantidad = float(cantidad) if cantidad else 0.0
        except:
            flash("Cantidad inválida", "danger")
            return redirect(url_for("kardex_list"))

        try:
            h_final = float(h_final) if h_final else None
        except:
            flash("Horómetro inválido", "danger")
            return redirect(url_for("kardex_list"))

        parte = request.form.get("parte_diario")
        obs = request.form.get("observacion")

        if tipo not in ["ENTRADA", "SALIDA", "OPERACION"]:
            flash("Tipo inválido", "danger")
            return redirect(url_for("kardex_list"))

        # CÓDIGO ACTUAL (Obsoleto):
        # tanque = Tanque.query.get(int(tanque_id))

        # CÓDIGO CORREGIDO (Moderno):
        tanque = db.session.get(Tanque, int(tanque_id)) if tanque_id else None


        # =========================
        # VALIDACIÓN SALIDA / OPERACION
        # =========================
        if tipo in ["SALIDA", "OPERACION"]:

            if not tanque:
                flash("Debe seleccionar un tanque", "danger")
                return redirect(url_for("kardex_list"))

            if not vehiculo_id:
                flash("Debe seleccionar un vehículo", "danger")
                return redirect(url_for("kardex_list"))

            if cantidad <= 0:
                flash("Cantidad debe ser mayor a 0", "danger")
                return redirect(url_for("kardex_list"))

            if tanque.stock_actual < cantidad:
                flash(
                    f"Stock insuficiente ({tanque.stock_actual:.2f} gls)",
                    "danger"
                )
                return redirect(url_for("kardex_list"))

        # =========================
        # 🔥 HORÓMETRO INICIAL (EDITABLE / RESPALDO AUTOMÁTICO)
        # =========================
        h_inicial = 0
        h_inicial_form = request.form.get("horometro_inicial")
        print(h_inicial_form)
        try:
            # Intentamos usar el valor editado por el usuario en el formulario
            h_inicial = float(h_inicial_form) if h_inicial_form else 0.0
        except ValueError:
            h_inicial = 0.0

        # Si el usuario no ingresó nada o es 0, buscamos el último de la base de datos como respaldo
        if h_inicial == 0 and vehiculo_id:
            vehiculo_id = int(vehiculo_id)

            ultimo = Kardex.query.filter(
                Kardex.vehiculo_id == vehiculo_id,
                Kardex.horometro_final != None,
                Kardex.activo == True
            ).order_by(Kardex.fecha.desc()).first()

            if ultimo:
                h_inicial = float(ultimo.horometro_final)

        # =========================
        # ⚠️ NORMALIZAR HORÓMETRO
        # =========================
        if tipo == "SALIDA":

            if h_final is None:
                # 🔥 CASO CLAVE: NO INGRESA HORÓMETRO
                h_final = h_inicial

            if h_final < h_inicial:
                flash("Horómetro final no puede ser menor", "danger")
                return redirect(url_for("kardex_list"))
        
        # Valida tipo entrada de combustible a tanque general
        if tipo == "ENTRADA":

            if not tanque:
                flash("Debe seleccionar un tanque para la compra", "danger")
                return redirect(url_for("kardex_list"))

            if cantidad <= 0:
                flash("Cantidad inválida", "danger")
                return redirect(url_for("kardex_list"))

            stock_resultante = tanque.stock_actual + cantidad

            if stock_resultante > tanque.capacidad:

                disponible = tanque.capacidad - tanque.stock_actual

                flash(
                    f"Capacidad excedida. "
                    f"Disponible: {disponible:.2f} gls. "
                    f"Capacidad máxima: {tanque.capacidad:.2f} gls.",
                    "danger"
                )

                return redirect(url_for("kardex_list"))
            
            if precio_unitario <= 0.0:
                flash(
                    "Debe ingresar precio por galón",
                    "danger"
                )

                return redirect(url_for("kardex_list"))
            
        costo_total = cantidad * precio_unitario

        # =========================
        # FECHA DEL MOVIMIENTO
        # =========================

        if fecha_str:
            fecha_movimiento = datetime.strptime(
                fecha_str,
                "%Y-%m-%dT%H:%M"
            )
        else:
            fecha_movimiento = now_lima()

        if fecha_movimiento > now_lima():
            flash(
                "No puede registrar movimientos futuros",
                "danger"
            )
            return redirect(
                url_for("kardex_list")
            )
        # =========================
        # RECORRIDO
        # =========================
        recorrido = h_final - h_inicial if h_final is not None else 0

        # =========================
        # CREAR MOVIMIENTO
        # =========================
        nuevo = Kardex(
            tipo=tipo,
            fecha=fecha_movimiento,
            proyecto_id=proyecto_id,
            tanque_id=int(tanque_id) if tanque_id else None,
            vehiculo_id=vehiculo_id,
            usuario_id=current_user.id,
            cantidad=cantidad,

            precio_unitario=precio_unitario,
            costo_total=costo_total,

            proveedor=proveedor,
            factura=factura,

            horometro_inicial=h_inicial,
            horometro_final=h_final,
            tanque_lleno=tanque_lleno,
            parte_diario=parte,
            observacion=obs,
            operador_id=int(operador_id) if operador_id else None
        )

        db.session.add(nuevo)
        db.session.flush()  # 🔥 IMPORTANTE

        # =========================
        # STOCK (CORREGIDO)
        # =========================
        if tipo == "ENTRADA":
            # 1. Calculamos el valor con el stock real ANTES de la compra
            tanque_precio_prom = float(tanque.precio_promedio) if tanque.precio_promedio else 0.0
            tanque_stock_previo = float(tanque.stock_actual)

            valor_stock_actual = tanque_stock_previo * tanque_precio_prom
            valor_compra = cantidad * precio_unitario

            # 2. RECIÉN AHORA actualizamos el stock físico en la base de datos
            tanque.stock_actual += cantidad
            nuevo_stock = float(tanque.stock_actual)

            # 3. Calculamos el promedio ponderado real
            if nuevo_stock > 0:
                tanque.precio_promedio = (valor_stock_actual + valor_compra) / nuevo_stock


        elif tipo in ["SALIDA", "OPERACION"]:

            if tanque:

                precio_unitario = float(
                    tanque.precio_promedio
                ) if tanque.precio_promedio else 0.0

                costo_total = cantidad * precio_unitario

                nuevo.precio_unitario = precio_unitario
                nuevo.costo_total = costo_total

                tanque.stock_actual -= cantidad        

        # =========================
        # 🚀 CONTROL TANQUE LLENO (CORREGIDO)
        # =========================
        if tipo == "SALIDA" and tanque_lleno and vehiculo_id:

            anterior = Kardex.query.filter(
                Kardex.vehiculo_id == vehiculo_id,
                Kardex.tipo == "SALIDA",
                Kardex.tanque_lleno == True,
                Kardex.activo == True,
                Kardex.fecha < nuevo.fecha
            ).order_by(Kardex.fecha.desc()).first()

            if anterior:
                h_ini = float(anterior.horometro_final) if anterior.horometro_final else 0.0
                h_fin = float(h_final) if h_final else 0.0

                recorrido_total = h_fin - h_ini

                # Capturamos la suma y la transformamos a float seguro inmediatamente
                consumo_scalar = db.session.query(func.sum(Kardex.cantidad)).filter(
                    Kardex.vehiculo_id == vehiculo_id,
                    Kardex.tipo == "SALIDA",
                    Kardex.activo == True,
                    Kardex.cantidad > 0,
                    Kardex.id > anterior.id,  # 🔥 Más preciso que filtrar solo por fecha
                    Kardex.fecha <= nuevo.fecha
                ).scalar()

                consumo_total = float(consumo_scalar) if consumo_scalar else 0.0

                # Evitamos división entre cero y errores de tipos mezclados
                if consumo_total > 0.0 and recorrido_total > 0.0:
                    
                    rendimiento = consumo_total / recorrido_total

                    # CORRECCIÓN: API Moderna de SQLAlchemy 2.0
                    vehiculo = db.session.get(Vehiculo, vehiculo_id)
                    estado = "NORMAL"

                    if vehiculo and vehiculo.rendimiento_promedio:
                        prom = float(vehiculo.rendimiento_promedio)

                        # rendimiento_calculado = galones por hora/KM
                        # Menor valor = mejor rendimiento (consume menos)
                        # Mayor valor = peor rendimiento (consume más)
                        if rendimiento > prom * 1.2:
                            estado = "BAJO"   # 🔴 Consume mucho
                        elif rendimiento < prom * 0.8:
                            estado = "ALTO"   # 🟢 Consume poco

                        if estado == "BAJO":
                            alerta = Alerta(
                                tipo="RENDIMIENTO_BAJO",
                                mensaje=f"Vehículo {vehiculo_id} bajo rendimiento: {rendimiento:.2f} Gls/H",
                                vehiculo_id=vehiculo_id
                            )
                            db.session.add(alerta)
            
                    nuevo_rend = Rendimiento(
                        vehiculo_id=vehiculo_id,
                        proyecto_id=nuevo.proyecto_id,
                        consumo_total=consumo_total,
                        recorrido_total=recorrido_total,
                        rendimiento_calculado=rendimiento,
                        rendimiento_referencia=(
                            float(vehiculo.rendimiento_promedio)
                            if vehiculo and vehiculo.rendimiento_promedio
                            else None
                        ),
                        estado=estado,
                        tipo_control="TANQUE_LLENO",
                        observacion="Control tanque lleno",
                        horometro_abastecimiento_inicial=h_ini,
                        horometro_abastecimiento_final=h_fin
                    )

                    db.session.add(nuevo_rend)

        # Confirmación final de la transacción en la BD
        db.session.commit()
        flash("✅ Movimiento registrado con éxito", "success")


    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", "danger")

    return redirect(url_for("kardex_list"))

@app.route("/kardex/anular/<int:id>", methods=["POST"])
@login_required
@permission_required("kardex", "eliminar")
def kardex_anular(id):

    try:

        mov = Kardex.query.get_or_404(id)

        # ===================================
        # YA ANULADO
        # ===================================
        if not mov.activo:

            flash(
                "Movimiento ya fue anulado",
                "warning"
            )

            return redirect(
                url_for("kardex_list")
            )

        tanque = mov.tanque

        # CORRECCIÓN: API Moderna de SQLAlchemy 2.0 (Reemplaza la línea vieja del principio)
        mov = db.session.get(Kardex, id) or abort(404)

        if not mov.activo:
            flash("Movimiento ya fue anulado", "warning")
            return redirect(url_for("kardex_list"))

        tanque = mov.tanque

        # ===================================
        # REVERSAR STOCK Y PRECIO PROMEDIO
        # ===================================
        if mov.tipo == "ENTRADA":
            # 1. Aseguramos tipos float para evitar colisiones NoneType / Decimal
            tanque_precio_prom = float(tanque.precio_promedio) if tanque.precio_promedio else 0.0
            tanque_stock_actual = float(tanque.stock_actual)
            
            mov_cantidad = float(mov.cantidad)
            mov_precio = float(mov.precio_unitario) if mov.precio_unitario else 0.0

            nuevo_stock = tanque_stock_actual - mov_cantidad

            if nuevo_stock < 0:
                flash("No se puede anular porque dejaría stock negativo en el tanque.", "danger")
                return redirect(url_for("kardex_list"))

            # 2. Reversar la ponderación económica (Dinero actual menos dinero de la compra anulada)
            valor_inventario_actual = tanque_stock_actual * tanque_precio_prom
            valor_compra_anulada = mov_cantidad * mov_precio
            nuevo_valor_inventario = valor_inventario_actual - valor_compra_anulada

            # 3. Aplicar los nuevos valores calculados al tanque
            tanque.stock_actual = nuevo_stock
            
            if nuevo_stock > 0:
                tanque.precio_promedio = nuevo_valor_inventario / nuevo_stock
            else:
                tanque.precio_promedio = 0.00  # Si el tanque se queda vacío, el precio vuelve a cero

        elif mov.tipo in ["SALIDA", "OPERACION"]:

            # devolver combustible al tanque
            tanque.stock_actual += float(mov.cantidad)

        # ===================================
        # ELIMINAR RENDIMIENTOS GENERADOS
        # ===================================
        if (
            mov.tipo == "SALIDA"
            and hasattr(Rendimiento, "kardex_id")
        ):

            Rendimiento.query.filter(
                Rendimiento.kardex_id == mov.id
            ).delete(
                synchronize_session=False
            )

        # ===================================
        # ELIMINAR ALERTAS GENERADAS
        # ===================================
        if hasattr(Alerta, "kardex_id"):

            Alerta.query.filter(
                Alerta.kardex_id == mov.id
            ).delete(
                synchronize_session=False
            )

        # ===================================
        # ANULACIÓN LÓGICA
        # ===================================
        mov.activo = False
        mov.fecha_anulacion = now_lima()
        mov.anulado_por = current_user.id

        db.session.flush()

        # ===================================
        # RECALCULAR RENDIMIENTOS
        # ===================================
        if (
            mov.tipo == "SALIDA"
            and mov.vehiculo_id
        ):

            recalcular_rendimientos_vehiculo(
                mov.vehiculo_id
            )

        db.session.commit()

        flash(
            "✅ Movimiento anulado correctamente",
            "success"
        )

    except Exception as e:

        db.session.rollback()

        flash(
            f"Error al anular: {str(e)}",
            "danger"
        )

    return redirect(
        url_for("kardex_list")
    )

def recalcular_rendimientos_vehiculo(vehiculo_id):
    # ===================================
    # ELIMINAR CÁLCULOS ANTERIORES (Sintaxis 2.0 segura)
    # ===================================
    db.session.execute(
        db.delete(Rendimiento).where(Rendimiento.vehiculo_id == vehiculo_id)
    )
    db.session.execute(
        db.delete(Alerta).where(Alerta.vehiculo_id == vehiculo_id)
    )

    # ===================================
    # OBTENER ABASTECIMIENTOS ACTIVOS
    # ===================================
    abastecimientos = db.session.scalars(
        db.select(Kardex).where(
            Kardex.vehiculo_id == vehiculo_id,
            Kardex.tipo == "SALIDA",
            Kardex.tanque_lleno == True,
            Kardex.activo == True,
            Kardex.cantidad > 0
        ).order_by(Kardex.fecha.asc())
    ).all()

    # Necesitamos al menos dos puntos para calcular una diferencia/recorrido
    if len(abastecimientos) < 2:
        return

    # ===================================
    # PROCESAR CÁLCULOS POR PARES
    # ===================================
    for i in range(1, len(abastecimientos)):
        anterior = abastecimientos[i - 1]
        actual = abastecimientos[i]

        h_ini = float(anterior.horometro_final) if anterior.horometro_final else 0.0
        h_fin = float(actual.horometro_final) if actual.horometro_final else 0.0

        recorrido_total = h_fin - h_ini

        # Sumatoria limpia aislando tipos Decimal de la BD a float de Python
        consumo_scalar = db.session.query(func.sum(Kardex.cantidad)).filter(
            Kardex.vehiculo_id == vehiculo_id,
            Kardex.tipo == "SALIDA",
            Kardex.activo == True,
            Kardex.cantidad > 0,
            Kardex.id > anterior.id,  # 🔥 Cambiado a ID para evitar desajustes por milisegundos
            Kardex.fecha > anterior.fecha,
            Kardex.fecha <= actual.fecha
        ).scalar()

        consumo_total = float(consumo_scalar) if consumo_scalar else 0.0

        # Validaciones de seguridad matemática
        if consumo_total <= 0.0 or recorrido_total <= 0.0:
            continue

        rendimiento = consumo_total / recorrido_total

        # API Moderna de SQLAlchemy 2.0 para buscar el vehículo
        vehiculo = db.session.get(Vehiculo, vehiculo_id)
        estado = "NORMAL"

        if vehiculo and vehiculo.rendimiento_promedio:
            prom = float(vehiculo.rendimiento_promedio)

            if rendimiento > prom * 1.2:
                estado = "BAJO"  # Consumo excesivo (malo)
            elif rendimiento < prom * 0.8:
                estado = "ALTO"  # Consumo óptimo (bueno)

        # Crear el nuevo historial recalculado
        nuevo_rend = Rendimiento(
            vehiculo_id=vehiculo_id,
            proyecto_id=actual.proyecto_id,
            consumo_total=consumo_total,
            recorrido_total=recorrido_total,
            rendimiento_calculado=rendimiento,
            rendimiento_referencia=(
                float(vehiculo.rendimiento_promedio)
                if vehiculo and vehiculo.rendimiento_promedio
                else None
            ),
            estado=estado,
            tipo_control="TANQUE_LLENO",
            observacion="Recalculado automáticamente",
            horometro_abastecimiento_inicial=h_ini,
            horometro_abastecimiento_final=h_fin
        )
        db.session.add(nuevo_rend)

        if estado == "BAJO":
            alerta = Alerta(
                tipo="RENDIMIENTO_BAJO",
                mensaje=f"Vehículo {vehiculo_id} bajo rendimiento detectado en recálculo",
                vehiculo_id=vehiculo_id
            )
            db.session.add(alerta)

@app.route("/kardex/ultimo_horometro/<int:vehiculo_id>")
@login_required
def ultimo_horometro(vehiculo_id):

    ultimo = Kardex.query.filter_by(vehiculo_id=vehiculo_id,activo=True)\
        .order_by(Kardex.fecha.desc()).first()

    return {
        "horometro_final": ultimo.horometro_final if ultimo else 0
    }

# =========================
# RENDIMIENTOS DASH
# =========================
@app.route("/rendimientos")
@login_required
@permission_required("rendimientos", "ver")
def rendimiento_list():

    from sqlalchemy import func

    proyecto_id = request.args.get(
    "proyecto_id",
    type=int
    )
    proyectos = Proyecto.query.filter_by(
    activo=True
    ).all()

    # =========================
    # FILTRO BASE
    # =========================
    filtro = []

    if proyecto_id:
        filtro.append(
        Rendimiento.proyecto_id == proyecto_id
    )

    # =========================
    # KPIs GENERALES
    # =========================
    total = db.session.query(
        func.count(Rendimiento.id)
    ).filter(
        *filtro
    ).scalar() or 0

    promedio = db.session.query(
        func.avg(Rendimiento.rendimiento_calculado)
    ).filter(
        *filtro
    ).scalar() or 0

    total_consumo = db.session.query(
        func.sum(Rendimiento.consumo_total)
    ).filter(
        *filtro
    ).scalar() or 0

    total_recorrido = db.session.query(
        func.sum(Rendimiento.recorrido_total)
    ).filter(
        *filtro
    ).scalar() or 0

    bajos = db.session.query(
        func.count(Rendimiento.id)
    ).filter(
        Rendimiento.estado == "BAJO",
        *filtro
    ).scalar() or 0

    porcentaje_bajo = (
        bajos / total * 100
    ) if total > 0 else 0

    ultimo = (
        Rendimiento.query
        .filter(*filtro)
        .order_by(Rendimiento.fecha.desc())
        .first()
    )

    tanque_lleno = db.session.query(
        func.count(Rendimiento.id)
    ).filter(
        Rendimiento.tipo_control == "TANQUE_LLENO",
        *filtro
    ).scalar() or 0

    # =========================
    # AGRUPACIÓN POR VEHÍCULO
    # =========================
    por_vehiculo = (
        db.session.query(
            Vehiculo.nombre,
            func.avg(
                Rendimiento.rendimiento_calculado
            ),
            func.count(
                Rendimiento.id
            )
        )
        .join(
            Vehiculo,
            Vehiculo.id == Rendimiento.vehiculo_id
        )
        .filter(
            *filtro
        )
        .group_by(
            Vehiculo.nombre
        )
        .all()
    )

    # =========================
    # AGRUPACIÓN POR PROYECTO
    # =========================
    por_proyecto = (
    db.session.query(
        Proyecto.nombre,
        func.avg(
            Rendimiento.rendimiento_calculado
        )
    )
    .join(
        Rendimiento,
        Rendimiento.proyecto_id == Proyecto.id
    )
    .filter(
        *filtro
    )
    .group_by(
        Proyecto.nombre
    )
    .all()
    )

    return render_template(
        "rendimientos.html",
        lista=(Rendimiento.query.filter(*filtro).order_by(Rendimiento.fecha.desc()).all()),
        promedio=round(promedio, 2),
        total_consumo=round(total_consumo, 2),
        total_recorrido=round(total_recorrido, 2),
        porcentaje_bajo=round(porcentaje_bajo, 1),
        ultimo=ultimo.rendimiento_calculado if ultimo else 0,
        tanque_lleno=tanque_lleno,
        proyectos=proyectos,
        proyecto_id=proyecto_id,
        por_vehiculo=por_vehiculo,
        por_proyecto=por_proyecto
    )


@app.route("/cargas", methods=["POST"])
@permission_required("registrar","crear")
def registrar_carga():
    data = request.json

    cantidad = float(data["cantidad"])

    precio_unitario = float(
        data.get("precio_unitario", 0)
    )

    costo_total = cantidad * precio_unitario

    nueva = Kardex(
        tipo="SALIDA",
        fecha=datetime.utcnow(),
        cantidad=cantidad,
        precio_unitario=precio_unitario,
        costo_total=costo_total,
        tanque_id=data["tanque_id"],
        vehiculo_id=data["vehiculo_id"],
        usuario_id=current_user.id,
        horometro=data.get("horometro"),
        kilometraje=data.get("kilometraje")
    )

    db.session.add(nueva)


    # actualizar stock
    tanque = Tanque.query.get(
        data["tanque_id"]
    )

    tanque.stock_actual -= cantidad

    db.session.commit()

    return {
        "msg": "Carga registrada"
    }

@app.route("/reportes/rendimientos")
@login_required
@permission_required("rendimientos", "ver")
def vista_reporte_rendimiento():

    return render_template(
        "reporte_rendimientos.html"
    )

# ===================================================================================
# Reporte Rendimiento Excel
# ===================================================================================
@app.route("/reportes/rendimientos/excel")
@login_required
@permission_required("rendimientos", "ver")
def reporte_rendimiento():

    tipo = request.args.get("tipo")
    fecha_str = request.args.get("fecha")

    if not fecha_str:
        return "Debe seleccionar una fecha", 400

    fecha_base = datetime.strptime(
        fecha_str,
        "%Y-%m-%d"
    )

    # =========================
    # RANGO DE FECHAS
    # =========================

    if tipo == "semanal":

        inicio = fecha_base - timedelta(
            days=fecha_base.weekday()
        )

        fin = inicio + timedelta(days=6)

    elif tipo == "mensual":

        inicio = fecha_base.replace(day=1)

        if fecha_base.month == 12:

            fin = fecha_base.replace(
                year=fecha_base.year + 1,
                month=1,
                day=1
            ) - timedelta(days=1)

        else:

            fin = fecha_base.replace(
                month=fecha_base.month + 1,
                day=1
            ) - timedelta(days=1)

    elif tipo == "anual":

        inicio = fecha_base.replace(
            month=1,
            day=1
        )

        fin = fecha_base.replace(
            month=12,
            day=31
        )

    else:
        return "Tipo inválido", 400

    # =========================
    # VEHICULOS
    # =========================
    query = Vehiculo.query.filter(
        Vehiculo.activo == True
    )

    if current_user.proyecto_id:

        query = query.filter(
            Vehiculo.proyecto_id == current_user.proyecto_id
        )

    vehiculos = query.order_by(
        Vehiculo.nombre
    ).all()

    # =========================
    # LISTA DE DÍAS
    # =========================

    dias = []

    current = inicio

    while current <= fin:

        dias.append(current)

        current += timedelta(days=1)

    # =========================
    # CONSUMOS AGRUPADOS
    # (UNA SOLA CONSULTA)
    # =========================

    ids_vehiculos = [v.id for v in vehiculos]

    consumos_dict = {}

    if ids_vehiculos:

        consumos = db.session.query(
            Kardex.vehiculo_id,
            func.date(Kardex.fecha).label("dia"),
            func.sum(Kardex.cantidad).label("total")
        ).filter(
            Kardex.tipo == "SALIDA",
            Kardex.activo == True,
            Kardex.vehiculo_id.in_(ids_vehiculos),
            Kardex.fecha >= inicio,
            Kardex.fecha < fin + timedelta(days=1)
        ).group_by(
            Kardex.vehiculo_id,
            func.date(Kardex.fecha)
        ).all()

        for c in consumos:

            consumos_dict[
                (c.vehiculo_id, c.dia)
            ] = float(c.total or 0)
        

    # =========================
    # CREAR EXCEL
    # =========================

    wb = Workbook()

    ws = wb.active

    ws.title = "Reporte"

    # =========================
    # CABECERA
    # =========================

    ws["A1"] = "REPORTE DE CONSUMO DE COMBUSTIBLE"
    ws["A2"] = f"Periodo: {inicio.date()} - {fin.date()}"

    # =========================
    # ENCABEZADOS
    # =========================

    ws.cell(row=4, column=1, value="ITEM")
    ws.cell(row=4, column=2, value="EQUIPO")

    col_total = 3

    for d in dias:

        ws.cell(
            row=4,
            column=col_total,
            value=d.strftime("%d-%b")
        )

        col_total += 1

    ws.cell(
        row=4,
        column=col_total,
        value="TOTAL"
    )

    # Resaltar TOTAL

    fill_yellow = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid"
    )

    ws.cell(
        row=4,
        column=col_total
    ).fill = fill_yellow

    # =========================
    # DETALLE
    # =========================

    row = 5

    for i, vehiculo in enumerate(
        vehiculos,
        start=1
    ):
        

        ws.cell(
            row=row,
            column=1,
            value=i
        )

        ws.cell(
            row=row,
            column=2,
            value=vehiculo.nombre
        )

        total_consumo = 0

        col = 3

        for d in dias:

            consumo = consumos_dict.get(
                (vehiculo.id, d.date()),
                0
            )

            ws.cell(
                row=row,
                column=col,
                value=round(consumo, 2)
            )

            total_consumo += consumo

            col += 1

        ws.cell(
            row=row,
            column=col,
            value=round(total_consumo, 2)
        )

        row += 1

    # =========================
    # FORMATO
    # =========================

    ws.freeze_panes = "C5"

    for columna in ws.columns:

        max_length = 0

        try:

            letra = columna[0].column_letter

            for celda in columna:

                if celda.value:

                    max_length = max(
                        max_length,
                        len(str(celda.value))
                    )

            ws.column_dimensions[
                letra
            ].width = max_length + 2

        except:
            pass

    # =========================
    # EXPORTAR
    # =========================

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)
    

    return send_file(
        output,
        as_attachment=True,
        download_name=f"reporte_{tipo}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/dashboard")
@login_required
def dashboard():

    proyecto_id = request.args.get(
    "proyecto_id",
    type=int
    )
    proyectos = Proyecto.query.filter_by(
    activo=True
    ).all()

    # Filtros del dashboard
    filtro = [
        Kardex.activo == True
    ]

    if proyecto_id:
        filtro.append(
            Kardex.proyecto_id == proyecto_id
        )

    # 1. Lista de últimos movimientos (Sintaxis 2.0 limpia)
    lista = db.session.scalars(
        db.select(Kardex)
        .where(Kardex.tipo == "SALIDA", *filtro)
        .order_by(Kardex.fecha.desc())
        .limit(10) # Acotamos a los últimos 10 para no saturar el servidor
    ).all()

    # 2. Consumo Total de Combustible
    total_consumo_scalar = db.session.query(func.sum(Kardex.cantidad)).filter(
        Kardex.tipo == "SALIDA", *filtro
    ).scalar()
    total_consumo = float(total_consumo_scalar) if total_consumo_scalar else 0.0

    # 3. Total de Despachos (Conteo)
    total_abastecimientos = db.session.scalar(
        db.select(func.count(Kardex.id)).where(Kardex.tipo == "SALIDA", *filtro)
    ) or 0

    # 4. CORRECCIÓN: Conteo correcto de Vehículos Únicos Activos
    vehiculos_activos = db.session.scalar(
        db.select(func.count(Kardex.vehiculo_id.distinct())).where(
            Kardex.tipo == "SALIDA", *filtro
        )
    ) or 0

    # 5. Promedio por despacho
    consumo_promedio = round(total_consumo / total_abastecimientos, 2) if total_abastecimientos else 0.0

    # 6. CORRECCIÓN GRÁFICOS: Mapeo seguro convirtiendo Decimal a float
    por_vehiculo_raw = db.session.query(Vehiculo.nombre, func.sum(Kardex.cantidad))\
        .join(Kardex, Kardex.vehiculo_id == Vehiculo.id)\
        .filter(Kardex.tipo == "SALIDA", *filtro)\
        .group_by(Vehiculo.nombre).all()
    por_vehiculo = [[v[0], float(v[1])] for v in por_vehiculo_raw]

    por_operador_raw = db.session.query(Operador.nombre, func.sum(Kardex.cantidad))\
        .join(Kardex, Kardex.operador_id == Operador.id)\
        .filter(Kardex.tipo == "SALIDA", *filtro)\
        .group_by(Operador.nombre).all()
    por_operador = [[o[0], float(o[1])] for o in por_operador_raw]

    # 7. Información de tanques e inventario físico total
    total_tanques = db.session.scalar(db.select(func.count(Tanque.id))) or 0
    stock_total_scalar = db.session.query(func.sum(Tanque.stock_actual)).scalar()
    
    stock_total = float(stock_total_scalar) if stock_total_scalar else 0.0

    # ==================================================================
    # 🌟 NUEVOS ADITAMENTOS OPERATIVOS CRÍTICOS
    # ==================================================================
    # Alertas activas de rendimiento deficiente para el operador
    alertas_criticas = db.session.scalar(
        db.select(func.count(Alerta.id)).where(Alerta.tipo == "RENDIMIENTO_BAJO")
    ) or 0

    # Lista de tanques que están por debajo del 20% de su capacidad total
    tanques_criticos = db.session.scalars(
        db.select(Tanque).where(Tanque.stock_actual <= (Tanque.capacidad * 0.20))
    ).all()

    return render_template(
        "dashboard_operativo.html",
        lista=lista,
        total_consumo=round(total_consumo, 2),
        total_abastecimientos=total_abastecimientos,
        vehiculos_activos=vehiculos_activos,
        consumo_promedio=consumo_promedio,
        por_vehiculo=por_vehiculo,
        por_operador=por_operador,
        total_tanques=total_tanques,
        stock_total=round(stock_total, 2),
        alertas_criticas=alertas_criticas,
        tanques_criticos=tanques_criticos,
        proyectos=proyectos,
        proyecto_id=proyecto_id
    )

@app.route("/dashboard/gerencial")
@login_required
def dashboard_gerencial():

    proyecto_id = request.args.get(
        "proyecto_id",
        type=int
        )
    proyectos = Proyecto.query.filter_by(
        activo=True
        ).all()
    # Filtros del dashboard
    filtro = [
            Kardex.activo == True
        ]

    if proyecto_id:
            filtro.append(
                Kardex.proyecto_id == proyecto_id
            )
    # ==========================
    # KPI GENERALES
    # ==========================

    total_compras = db.session.query(
        func.sum(Kardex.cantidad)
    ).filter(
        Kardex.tipo == "ENTRADA",
        *filtro
    ).scalar() or 0

    total_consumo = db.session.query(
        func.sum(Kardex.cantidad)
    ).filter(
        Kardex.tipo == "SALIDA",
        *filtro
    ).scalar() or 0

    total_operaciones = Kardex.query.filter(
    *filtro).count()

    query_tanque = db.session.query(
        func.sum(Tanque.stock_actual)
    )

    if proyecto_id:

        query_tanque = query_tanque.filter(
            Tanque.proyecto_id == proyecto_id
        )

    stock_total = query_tanque.scalar() or 0

    total_tanques = Tanque.query.count()

    total_vehiculos = Vehiculo.query.count()

    balance = total_compras - total_consumo

    # ==========================
    # COMPRAS VS CONSUMO
    # ==========================
    compras = db.session.query(
    func.sum(Kardex.cantidad)
    ).filter(
        Kardex.tipo == "ENTRADA",
        *filtro
    ).scalar() or 0

    consumos = db.session.query(
    func.sum(Kardex.cantidad)
    ).filter(
        Kardex.tipo == "SALIDA",
        *filtro
    ).scalar() or 0

    # ==========================
    # TOP VEHICULOS
    # ==========================

    por_vehiculo = db.session.query(
        Vehiculo.nombre,
        func.sum(Kardex.cantidad)
    ).join(
        Kardex,
        Kardex.vehiculo_id == Vehiculo.id
    ).filter(
        Kardex.tipo == "SALIDA",
        *filtro
    ).group_by(
        Vehiculo.nombre
    ).all()

    # ==========================
    # TOP OPERADORES
    # ==========================

    por_operador = db.session.query(
        Operador.nombre,
        func.sum(Kardex.cantidad)
    ).join(
        Kardex,
        Kardex.operador_id == Operador.id
    ).filter(
        Kardex.tipo == "SALIDA",
        *filtro
    ).group_by(
        Operador.nombre
    ).all()

    # ==========================
    # ULTIMOS MOVIMIENTOS
    # ==========================

    movimientos = Kardex.query.filter(
    *filtro).order_by(
        Kardex.fecha.desc()
    ).limit(20).all()

    # ==================================================================
    # 💰 NUEVO: VALORIZACIONES ECONÓMICAS (GERENCIALES)
    # ==================================================================
    # Suma total de costo_total invertido/gastado en las salidas de combustible activas
    valor_consumo_scalar = db.session.query(func.sum(Kardex.costo_total)).filter(
        Kardex.tipo == "SALIDA", *filtro
    ).scalar()
    valor_total_consumo = float(valor_consumo_scalar) if valor_consumo_scalar else 0.0

    # Valorización monetaria del inventario actual en tanques (Stock * Precio Promedio Ponderado)
    # Usamos db.session.query para hacer la matemática directa y veloz en la base de datos
    valor_inventario_scalar = db.session.query(
        func.sum(Tanque.stock_actual * Tanque.precio_promedio)
    ).scalar()
    valor_total_inventario = float(valor_inventario_scalar) if valor_inventario_scalar else 0.0

    print(movimientos)
    
    return render_template(
        "dashboard_gerencial.html",

        total_compras=round(total_compras,2),
        total_consumo=round(total_consumo,2),
        stock_total=round(stock_total,2),
        balance=round(balance,2),

        total_tanques=total_tanques,
        total_vehiculos=total_vehiculos,
        total_operaciones=total_operaciones,

        compras=compras,
        consumos=consumos,

        por_vehiculo=por_vehiculo,
        por_operador=por_operador,

        movimientos=movimientos,
        # Variables financieras añadidas:
        valor_total_consumo=round(valor_total_consumo, 2),
        valor_total_inventario=round(valor_total_inventario, 2),
        proyectos=proyectos,
        proyecto_id=proyecto_id,
    )

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8007, debug=True)